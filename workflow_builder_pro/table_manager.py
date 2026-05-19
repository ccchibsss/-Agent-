"""
Менеджер для работы с таблицами (Google Sheets, Excel).
Поддерживает .xlsx, .xlsm, .xls (через xlrd), .csv.
Добавлены:
- max_rows для быстрой предзагрузки Excel
- запись в Google Sheets через gspread
"""
import streamlit as st
import pandas as pd
import requests
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
from openai import OpenAI
from io import BytesIO
import logging
import re
import json
from config import CONFIG
from utils import handle_errors, cache_result
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    openpyxl = None

try:
    import xlrd
    XLRD_SUPPORT = True
except ImportError:
    XLRD_SUPPORT = False
    xlrd = None

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSHEETS_WRITE_SUPPORT = True
except ImportError:
    GSHEETS_WRITE_SUPPORT = False
    gspread = None

logger = logging.getLogger(__name__)


class TableManager:
    def __init__(self, api_key: Optional[str] = None):
        self.api_key = api_key
        self._cache: Dict[str, pd.DataFrame] = {}
        self._last_operation: Optional[Dict] = None

    @handle_errors(default_return=None)
    def read_google_sheets(self, url: str, sheet_name: Optional[str] = None,
                           range_a1: Optional[str] = None, use_cache: bool = True) -> Optional[pd.DataFrame]:
        if '/d/' in url:
            sheet_id = url.split('/d/')[1].split('/')[0]
        else:
            sheet_id = url
        cache_key = f"gsheets:{sheet_id}:{sheet_name}:{range_a1}"
        if use_cache and cache_key in self._cache:
            return self._cache[cache_key].copy()
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export"
        params = {'format': 'csv'}
        if sheet_name:
            params['gid'] = self._get_sheet_gid(url, sheet_name)
        if range_a1:
            params['range'] = range_a1
        response = requests.get(csv_url, params=params, timeout=CONFIG.API_TIMEOUT)
        response.raise_for_status()
        df = pd.read_csv(BytesIO(response.content))
        if use_cache:
            self._cache[cache_key] = df.copy()
        self._last_operation = {
            'type': 'read', 'source': 'google_sheets', 'rows': len(df),
            'columns': list(df.columns), 'timestamp': datetime.now().isoformat()
        }
        return df

    def _get_sheet_gid(self, url: str, sheet_name: str) -> Optional[str]:
        return None

    @handle_errors(default_return=None)
    def read_excel(self, file_path: Union[str, BytesIO], sheet_name: Optional[Union[str, int]] = 0,
                   range_a1: Optional[str] = None, use_cache: bool = True,
                   max_rows: Optional[int] = None) -> Optional[pd.DataFrame]:
        """Читает Excel, можно ограничить число строк для ускорения предпросмотра."""
        if not EXCEL_SUPPORT:
            raise ImportError("Установите openpyxl: pip install openpyxl")

        ext = ""
        if isinstance(file_path, str):
            ext = os.path.splitext(file_path)[1].lower()
        elif isinstance(file_path, BytesIO):
            name = getattr(file_path, 'name', '')
            ext = os.path.splitext(name)[1].lower() if name else ''

        # Параметры для pd.read_excel
        kwargs = {'sheet_name': sheet_name}
        if range_a1:
            kwargs['usecols'] = range_a1
        if max_rows is not None:
            kwargs['nrows'] = max_rows

        try:
            df = pd.read_excel(file_path, engine='openpyxl', **kwargs)
        except Exception as e:
            if XLRD_SUPPORT and ext == '.xls':
                logger.info("Попытка чтения .xls через xlrd")
                df = pd.read_excel(file_path, engine='xlrd', **kwargs)
            else:
                logger.error(f"Ошибка чтения Excel: {e}")
                return None

        if use_cache:
            cache_key = f"excel:{hash(str(file_path))}:{sheet_name}:{max_rows}"
            self._cache[cache_key] = df.copy()
        self._last_operation = {
            'type': 'read', 'source': 'excel', 'rows': len(df),
            'columns': list(df.columns), 'timestamp': datetime.now().isoformat()
        }
        return df

    @handle_errors(default_return=False)
    def write_excel(self, df: pd.DataFrame, output_path: str, sheet_name: str = 'Sheet1',
                    apply_formatting: bool = True, formatting_rules: Optional[Dict] = None) -> bool:
        if not EXCEL_SUPPORT:
            raise ImportError("Требуется openpyxl")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            if apply_formatting:
                self._apply_excel_formatting(writer, df, formatting_rules)
        return True

    def _apply_excel_formatting(self, writer, df, rules):
        worksheet = writer.sheets[writer.sheet_names[0]]
        for column in worksheet.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column)
            col_letter = column[0].column_letter
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        if rules:
            self._apply_custom_rules(worksheet, df, rules)

    def _apply_custom_rules(self, worksheet, df, rules):
        if rules.get('conditional_format'):
            for rule in rules['conditional_format']:
                col_name = rule.get('column')
                condition = rule.get('condition')
                format_config = rule.get('format', {})
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    self._apply_conditional_format(worksheet, col_idx, len(df) + 1, condition, format_config)

    def _apply_conditional_format(self, worksheet, col_idx, max_row, condition, format_config):
        fill_color = format_config.get('color', 'FFEB3B')
        for row in range(2, max_row + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            value = cell.value
            if condition == 'highlight_negatives' and isinstance(value, (int, float)) and value < 0:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            elif condition == 'highlight_high' and isinstance(value, (int, float)) and value > 1000:
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type="solid")

    @cache_result()
    def ai_analyze_dataframe(self, df: pd.DataFrame, instruction: str, api_key: str) -> Dict[str, Any]:
        if not api_key:
            return {'error': 'API ключ не указан'}
        df_summary = {
            'shape': df.shape,
            'columns': list(df.columns),
            'dtypes': {str(k): str(v) for k, v in df.dtypes.to_dict().items()},
            'sample': df.head(3).to_dict('records'),
            'null_counts': df.isnull().sum().to_dict(),
            'stats': df.describe(include='all').to_dict() if not df.empty else {}
        }
        client = OpenAI(api_key=api_key, base_url=CONFIG.DEEPSEEK_BASE_URL)
        prompt = f"""
Ты эксперт по анализу данных в Python pandas.

ДАННЫЕ:
{json.dumps(df_summary, ensure_ascii=False, default=str)[:8000]}

ЗАДАЧА: {instruction}

Верни ответ в формате JSON:
{{
    "analysis": "Краткий анализ данных на русском",
    "issues_found": ["список проблем"],
    "recommendations": ["список рекомендаций"],
    "transformations": [
        {{
            "type": "clean|filter|aggregate|pivot|format",
            "code": "pandas код для выполнения",
            "description": "что делает этот код"
        }}
    ],
    "ready_code": "полный готовый код для копирования"
}}

Отвечай ТОЛЬКО валидным JSON, без пояснений.
"""
        try:
            response = client.chat.completions.create(
                model=CONFIG.DEEPSEEK_MODEL,
                messages=[{"role": "system", "content": "Ты аналитик данных. Отвечай только валидным JSON."},
                          {"role": "user", "content": prompt}],
                temperature=0.2, timeout=CONFIG.API_TIMEOUT, max_tokens=CONFIG.MAX_TOKENS
            )
            content = response.choices[0].message.content
            if "```json" in content:
                content = content.split("```json", 1)[1].split("```", 1)[0]
            elif "```" in content:
                content = content.split("```", 1)[1].split("```", 1)[0]
            json_match = re.search(r'\{[\s\S]*\}', content)
            if json_match:
                json_str = json_match.group()
                json_str = re.sub(r'//[^\n]*', '', json_str)
                json_str = re.sub(r'/\*[\s\S]*?\*/', '', json_str)
                json_str = re.sub(r',(\s*[\]}])', r'\1', json_str)
                try:
                    return json.loads(json_str)
                except json.JSONDecodeError as e:
                    return {'error': f'Ошибка парсинга JSON: {str(e)}'}
            return {'error': 'Не удалось извлечь JSON из ответа'}
        except Exception as e:
            return {'error': f'Ошибка ИИ: {str(e)}'}

    def execute_transformation(self, df: pd.DataFrame, transformation_code: str) -> pd.DataFrame:
        safe_globals = {
            "pd": pd,
            "np": __import__('numpy') if 'numpy' in transformation_code else None,
            "df": df.copy()
        }
        try:
            exec(transformation_code, safe_globals, safe_globals)
            return safe_globals.get('df', df)
        except Exception as e:
            logger.error(f"Ошибка выполнения: {e}")
            raise

    def write_google_sheets(self, df: pd.DataFrame, url: str, sheet_name: str = 'Sheet1',
                            mode: str = 'overwrite') -> bool:
        """
        Записывает DataFrame в Google Sheets.
        Требует настроенных секретов Streamlit: GSPREAD_CREDENTIALS (словарь ключей сервисного аккаунта).
        mode: 'overwrite' - замена всего листа, 'append' - добавление в конец (пока не реализовано).
        """
        if not GSHEETS_WRITE_SUPPORT:
            raise ImportError("Установите gspread: pip install gspread")

        creds_dict = None
        # Пробуем загрузить из secrets
        try:
            creds_dict = st.secrets["GSPREAD_CREDENTIALS"]
            if isinstance(creds_dict, str):
                creds_dict = json.loads(creds_dict)
        except KeyError:
            pass

        if not creds_dict:
            raise PermissionError("Не настроены GSPREAD_CREDENTIALS в секретах приложения.")

        try:
            credentials = Credentials.from_service_account_info(creds_dict)
            client = gspread.authorize(credentials)
            # Извлекаем ID таблицы из URL
            sheet_id = url.split('/d/')[1].split('/')[0] if '/d/' in url else url
            spreadsheet = client.open_by_key(sheet_id)
            try:
                worksheet = spreadsheet.worksheet(sheet_name)
            except gspread.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(title=sheet_name, rows="100", cols="20")

            # Очищаем лист и вставляем данные
            worksheet.clear()
            # Подготавливаем список списков
            data = [df.columns.tolist()] + df.values.tolist()
            worksheet.update(data)
            logger.info(f"Данные успешно записаны в Google Sheets: {sheet_name}")
            return True
        except Exception as e:
            logger.error(f"Ошибка записи в Google Sheets: {e}")
            raise
