import streamlit as st
import json
import time
import logging
from typing import Dict, List, Tuple, Any, Optional, Callable
from dataclasses import dataclass, field
from enum import Enum, auto
from pathlib import Path
from io import BytesIO
import base64
import pandas as pd
from config import (
    WORKFLOW_FILE, AGENTS_FILE, MESSAGES_FILE, HISTORY_FILE,
    TABLES_FILE, IMAGES_METADATA_FILE, IMAGES_DIR, CONFIG
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)


def save_workflow_auto(workflow: List[Dict]):
    try:
        with open(WORKFLOW_FILE, 'w', encoding='utf-8') as f:
            json.dump(workflow, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Не удалось сохранить workflow: {e}")


def load_workflow_auto() -> List[Dict]:
    if WORKFLOW_FILE.exists():
        try:
            with open(WORKFLOW_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Не удалось загрузить workflow: {e}")
    return []


def save_agents_auto(agents_data: Dict):
    try:
        with open(AGENTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(agents_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Не удалось сохранить агентов: {e}")


def load_agents_auto() -> Optional[Dict]:
    if AGENTS_FILE.exists():
        try:
            with open(AGENTS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Не удалось загрузить агентов: {e}")
    return None


def save_messages_auto(messages: List[Dict]):
    try:
        with open(MESSAGES_FILE, 'w', encoding='utf-8') as f:
            json.dump(messages, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Не удалось сохранить сообщения: {e}")


def load_messages_auto() -> List[Dict]:
    if MESSAGES_FILE.exists():
        try:
            with open(MESSAGES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Не удалось загрузить сообщения: {e}")
    return []


def save_history_auto(history: List[Dict]):
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Не удалось сохранить историю: {e}")


def load_history_auto() -> List[Dict]:
    if HISTORY_FILE.exists():
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Не удалось загрузить историю: {e}")
    return []


def save_tables_auto(tables_data: Dict):
    try:
        with open(TABLES_FILE, 'w', encoding='utf-8') as f:
            serializable = {}
            for key, value in tables_data.items():
                if isinstance(value, pd.DataFrame):
                    serializable[key] = {
                        'data': value.to_dict('records'),
                        'columns': list(value.columns),
                        'created_at': time.strftime('%Y-%m-%dT%H:%M:%S')
                    }
                else:
                    serializable[key] = value
            json.dump(serializable, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Не удалось сохранить таблицы: {e}")


def load_tables_auto() -> Dict:
    if TABLES_FILE.exists():
        try:
            with open(TABLES_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                result = {}
                for key, value in data.items():
                    if isinstance(value, dict) and 'data' in value:
                        result[key] = pd.DataFrame(value['data'])
                    else:
                        result[key] = value
                return result
        except Exception as e:
            logger.warning(f"Не удалось загрузить таблицы: {e}")
    return {}


def save_images_metadata_auto(metadata: Dict):
    try:
        with open(IMAGES_METADATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Не удалось сохранить метаданные изображений: {e}")


def load_images_metadata_auto() -> Dict:
    if IMAGES_METADATA_FILE.exists():
        try:
            with open(IMAGES_METADATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Не удалось загрузить метаданные изображений: {e}")
    return {}


def cache_result(ttl_seconds: int = CONFIG.CACHE_TTL_SECONDS):
    def decorator(func: Callable):
        cache: Dict[str, Tuple[Any, float]] = {}
        def wrapper(*args, **kwargs) -> Any:
            key = f"{func.__name__}:{hash(str(args) + str(sorted(kwargs.items())))}"
            current_time = time.time()
            if key in cache:
                result, timestamp = cache[key]
                if current_time - timestamp < ttl_seconds:
                    return result
            result = func(*args, **kwargs)
            cache[key] = (result, current_time)
            return result
        return wrapper
    return decorator


def handle_errors(default_return: Any = None):
    def decorator(func: Callable):
        def wrapper(*args, **kwargs) -> Any:
            try:
                return func(*args, **kwargs)
            except Exception as e:
                logger.error(f"Ошибка в {func.__name__}: {str(e)}", exc_info=True)
                st.error(f"⚠️ {func.__name__}: {str(e)}")
                return default_return
        return wrapper
    return decorator


def format_bytes(size: int) -> str:
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size < 1024.0:
            return f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{size:.1f} TB"


def image_to_base64(image) -> str:
    from PIL import Image
    buffered = BytesIO()
    image.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode()


def base64_to_image(base64_string: str):
    from PIL import Image
    image_data = base64.b64decode(base64_string)
    return Image.open(BytesIO(image_data))


class NodeType(Enum):
    GOOGLE_SHEETS_READ = "google_sheets_read"
    GOOGLE_SHEETS_WRITE = "google_sheets_write"
    EXCEL_READ = "excel_read"
    EXCEL_WRITE = "excel_write"
    EXCEL_FORMAT = "excel_format"
    EXCEL_CHART = "excel_chart"
    DEEPSEEK_AI = "deepseek"
    CONDITION = "condition"
    LOOP = "loop"
    HTTP_GET = "http_get"
    HTTP_POST = "http_post"
    EMAIL = "email"
    TELEGRAM = "telegram"
    AI_AGENT = "ai_agent"
    DATA_CLEAN = "data_clean"
    PIVOT_TABLE = "pivot_table"
    FILTER = "filter"
    TRANSFORM = "transform"


class ConditionType(Enum):
    GREATER = "greater"
    LESS = "less"
    EQUAL = "equal"
    NOT_EQUAL = "not_equal"
    CONTAINS = "contains"
    NOT_CONTAINS = "not_contains"
    STARTS_WITH = "starts_with"
    ENDS_WITH = "ends_with"
    IS_EMPTY = "is_empty"
    IS_NOT_EMPTY = "is_not_empty"
    BETWEEN = "between"
    IN_LIST = "in_list"
    CUSTOM = "custom"


class MemoryImportance(Enum):
    LOW = "low"
    NORMAL = "normal"
    HIGH = "high"


class WorkflowStatus(Enum):
    PENDING = "pending"
    RUNNING = "running"
    SUCCESS = "success"
    ERROR = "error"
    PAUSED = "paused"


class ImageEditOperation(Enum):
    REMOVE_BACKGROUND = "remove_background"
    REMOVE_WATERMARK = "remove_watermark"
    RESIZE = "resize"
    CROP = "crop"
    ROTATE = "rotate"
    ENHANCE = "enhance"
    FILTER = "filter"
    ADD_TEXT = "add_text"
    ADD_WATERMARK = "add_watermark"
    CONVERT_FORMAT = "convert_format"


@dataclass
class TableCell:
    row: int
    column: int
    value: Any
    formula: Optional[str] = None
    font: Optional[Dict] = None
    fill: Optional[Dict] = None
    border: Optional[Dict] = None
    alignment: Optional[Dict] = None
    data_validation: Optional[Dict] = None
    
    def to_openpyxl_style(self) -> Dict[str, Any]:
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        styles = {}
        if self.font and Font:
            styles['font'] = Font(**{k: v for k, v in self.font.items() if v is not None})
        if self.fill and PatternFill:
            styles['fill'] = PatternFill(**self.fill)
        if self.border and Border:
            side = Side(**{k: v for k, v in self.border.items() if v is not None})
            styles['border'] = Border(left=side, right=side, top=side, bottom=side)
        if self.alignment and Alignment:
            styles['alignment'] = Alignment(**self.alignment)
        return styles


@dataclass
class TableRange:
    sheet_name: str
    start_row: int
    end_row: int
    start_col: Any
    end_col: Any
    
    @property
    def a1_notation(self) -> str:
        from openpyxl.utils import get_column_letter
        start_col_letter = get_column_letter(self.start_col) if isinstance(self.start_col, int) else self.start_col
        end_col_letter = get_column_letter(self.end_col) if isinstance(self.end_col, int) else self.end_col
        return f"{self.sheet_name}!{start_col_letter}{self.start_row}:{end_col_letter}{self.end_row}"


@dataclass
class ChartConfig:
    chart_type: str
    title: str
    x_column: str
    y_columns: List[str]
    x_title: Optional[str] = None
    y_title: Optional[str] = None
    colors: Optional[List[str]] = None
    width: int = 800
    height: int = 400
